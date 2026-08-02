#!/usr/bin/env python3
"""从结构化 JSON 用例数据生成研发自测用例 Excel（.xlsx）。

输入 JSON 的形态：
  1. {"cases": [ {用例对象}, ... ], "coverage": {覆盖核算对象(可选)} }
  2. [ {用例对象}, ... ]

用例对象字段（与 SKILL.md 的列一一对应）：
  case_id         用例编号，如 TC-ORD-001
  scenario        业务场景，如 订单管理
  function        功能，如 创建订单
  name            用例名称，如 创建订单-用户为空校验
  precondition    前置条件
  steps           操作步骤（list 或单行字符串）
  priority        优先级，P0/P1/P2
  expected_result 预期结果（不通过分支写功能展示的结果）
  test_result     测试结果（留空，列内下拉 PASS/FAIL）
  remark          备注
  code_location   代码位置，如 OrderService.java:3

coverage 对象（SKILL.md 第 4 步覆盖率核对产出，用于生成「覆盖说明」sheet）：
  stats        {"entry_covered":3, "entry_total":3, "dp_covered":22, "dp_total":22}
  inventory    [ {"function":"创建订单","decision_point":"if (...)","line":"OrderService.java:3","branches":["真→TC-ORD-001","假→TC-ORD-002"]}, ... ]
  uncovered    [ 缺口描述, ... ]（无缺口可为空数组）
  notes        [ 无法静态验证项, ... ]

用法：
  python generate_excel.py cases.json [输出路径.xlsx]
"""

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("缺少 openpyxl，请先执行: pip install openpyxl")

COLUMNS = [
    ("case_id", "用例编号"),
    ("scenario", "业务场景"),
    ("function", "功能"),
    ("name", "用例名称"),
    ("scenario_type", "场景类型"),
    ("verify_point", "验证点"),
    ("precondition", "前置条件"),
    ("steps", "操作步骤"),
    ("verify_method", "验证方法"),
    ("auto_level", "可自动化程度"),
    ("priority", "优先级"),
    ("expected_result", "预期结果"),
    ("test_result", "测试结果"),
    ("remark", "备注"),
    ("code_location", "代码位置"),
]

# 场景类型文字着色（v1.21）：正常流程整行绿色、异常流程整行红色，便于目视区分成功/拒绝路径
SCENARIO_TYPE_COLORS = {
    "正常流程": "008000",  # 绿
    "异常流程": "FF0000",  # 红
}

# 每列宽度（近似，按最长表头/内容预估）
COLUMN_WIDTHS = {
    "case_id": 14, "scenario": 16, "function": 16, "name": 28, "scenario_type": 12,
    "verify_point": 30, "precondition": 34, "steps": 42,
    "verify_method": 13, "auto_level": 15, "priority": 8,
    "expected_result": 44, "test_result": 12, "remark": 30, "code_location": 18,
}

# ===== v1.22 样式常量：统一字体 + 行高自适应 + 斑马纹 + 优先级着色 =====
FONT_NAME = "微软雅黑"            # 中文字体（Excel 渲染友好，跨平台回退）
FONT_SIZE = 10
HEADER_FONT_SIZE = 11
HEADER_FILL = PatternFill("solid", fgColor="4472C4")       # 表头蓝（主表 + 覆盖说明子表）
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")      # 分节标题浅蓝
BAND_FILL_EVEN = PatternFill("solid", fgColor="F2F7FB")    # 数据行斑马纹（偶数行浅蓝）
BAND_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")     # 数据行斑马纹（奇数行白）
PRIORITY_COLORS = {"P0": "C00000", "P1": "BF8F00", "P2": "7F7F7F"}  # 优先级文字：P0 红 / P1 琥珀 / P2 灰
SECTION_TITLE_COLOR = "1F4E79"   # 分节标题深蓝
HEADER_ROW_HEIGHT = 28           # 表头行高（pt）
LINE_HEIGHT = 16                 # 每行文字高度（pt）
MIN_ROW_HEIGHT = 22              # 数据行最小高度（pt）

# 通用边框 / 对齐（模块级复用）
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(vertical="center", horizontal="center")

# 行高估算用的 列索引 → 列宽 映射
MAIN_WIDTH_MAP = {i: COLUMN_WIDTHS.get(key, 20) for i, (key, _) in enumerate(COLUMNS, start=1)}
COVERAGE_WIDTH_MAP = {1: 28, 2: 32, 3: 20, 4: 50}


def _text_width(text):
    """计算文本显示宽度：中文/全角字符占 2 个单位，其余占 1（近似 Excel 列宽单位）。"""
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)


def _chars_per_line(width):
    """估算单行可容纳的显示宽度单位（列宽即宽度单位，留 1 单位余量防裁切）。"""
    return max(int(width) - 1, 4)


def _text_lines(text, width):
    """估算文本按列宽换行后的行数（含显式换行分隔）。按显示宽度计算，贴合实际换行。"""
    text = "" if text is None else str(text)
    if not text:
        return 1
    cpl = _chars_per_line(width)
    return sum(max(1, -(-_text_width(seg) // cpl)) for seg in text.split("\n"))


def fit_row_height(ws, row, width_map):
    """按行内容估算并设置行高，保证换行内容完全显示。width_map: {列索引: 列宽}。"""
    max_lines = 1
    for col_idx, width in width_map.items():
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is None:
            continue
        max_lines = max(max_lines, _text_lines(cell.value, width))
    ws.row_dimensions[row].height = max(MIN_ROW_HEIGHT, max_lines * LINE_HEIGHT + 4)


def load_data(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_section_title(ws, row, text, size=HEADER_FONT_SIZE):
    """分节标题：浅蓝填充 + 加粗深蓝文字 + 自动换行。"""
    c = ws.cell(row=row, column=1, value=text)
    c.fill = SECTION_FILL
    c.font = Font(name=FONT_NAME, size=size, bold=True, color=SECTION_TITLE_COLOR)
    c.alignment = WRAP_TOP
    return row


def write_table_header(ws, row, headers, start=1):
    """子表表头：蓝底白字加粗（与主 sheet 表头一致）。"""
    for col, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")
        c.alignment = CENTER
        c.border = BORDER
    return row


def build_coverage_sheet(wb, coverage, cases=None):
    """在同一工作簿新增「覆盖说明」sheet：覆盖统计（含变更范围）+ 追溯矩阵 + 标注区（缺口/无法验证/规则缺口）。"""
    ws = wb.create_sheet("覆盖说明")

    def cset(row, col, value):
        """写单元格：统一字体 + 边框 + 自动换行。"""
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
        cell.border = BORDER
        cell.alignment = WRAP_TOP
        return cell

    row = 1
    write_section_title(ws, row, "覆盖说明（代码可表达范围内）", size=12)
    row += 2

    stats = coverage.get("stats", {}) if isinstance(coverage, dict) else {}
    write_section_title(ws, row, "覆盖统计")
    row += 1
    for label, covered, total in (
        ("业务入口覆盖", stats.get("entry_covered"), stats.get("entry_total")),
        ("判定点覆盖（机器基准）", stats.get("dp_covered"), stats.get("dp_total")),
        ("需求验证点覆盖", stats.get("verification_covered"), stats.get("verification_total")),
    ):
        cset(row, 1, label)
        val = f"{covered}/{total}" if covered is not None and total is not None else ""
        cset(row, 2, val)
        row += 1

    # 可自动化程度分布（验证计划编排依据）
    auto_dist = (coverage or {}).get("auto_level_distribution")
    if auto_dist:
        row += 1
        write_section_title(ws, row, "可自动化程度分布")
        row += 1
        for level, cnt in auto_dist.items():
            cset(row, 1, str(level))
            cset(row, 2, str(cnt))
            row += 1

    # 场景类型分布（按功能统计正常/异常流程用例数）
    if cases:
        by_type = {}
        for c in cases:
            fn = c.get("function") or "?"
            by_type.setdefault(fn, {"正常流程": 0, "异常流程": 0})
            st = c.get("scenario_type")
            if st in by_type[fn]:
                by_type[fn][st] += 1
        if by_type:
            row += 1
            write_section_title(ws, row, "场景类型分布（按功能统计正常/异常流程用例数）")
            row += 1
            write_table_header(ws, row, ["功能", "正常流程", "异常流程"])
            row += 1
            for fn in sorted(by_type):
                cset(row, 1, fn)
                cset(row, 2, by_type[fn]["正常流程"])
                cset(row, 3, by_type[fn]["异常流程"])
                row += 1

    # 功能覆盖清单（研发视角：每个功能必须有 ≥1 条用例）
    per_function = (coverage or {}).get("per_function") if isinstance(coverage, dict) else None
    if per_function:
        row += 1
        write_section_title(ws, row, "功能覆盖清单（每个功能必须有 ≥1 条用例）")
        row += 1
        write_table_header(ws, row, ["功能", "判定点覆盖"])
        row += 1
        for fn, st in per_function.items():
            cset(row, 1, fn)
            cset(row, 2, f"{st.get('dp_covered', 0)}/{st.get('dp_total', 0)}")
            row += 1

    row += 1
    write_section_title(ws, row, "变更范围")
    row += 1
    change_scope = (coverage or {}).get("change_scope") if isinstance(coverage, dict) else None
    if change_scope:
        for key, label in (
            ("mode", "模式"),
            ("file_count", "变更文件数"),
            ("function_count", "受影响函数数"),
            ("affected_decision_points_total", "受影响函数内判定点总数"),
        ):
            if change_scope.get(key) is not None:
                cset(row, 1, label)
                cset(row, 2, change_scope.get(key))
                row += 1
        for key, label in (
            ("changed_files", "变更文件"),
            ("affected_functions", "受影响函数"),
            ("new_decision_points", "新增判定点"),
        ):
            if change_scope.get(key):
                cset(row, 1, label)
                cset(row, 2, "\n".join(change_scope[key]))
                row += 1

    row += 1
    write_section_title(ws, row, "判定点 → 用例 追溯矩阵")
    row += 1
    write_table_header(ws, row, ["功能", "判定点", "代码位置", "分支 → 用例编号"])
    row += 1
    for item in (coverage or {}).get("inventory") or []:
        branches = item.get("branches") or []
        if isinstance(branches, list):
            branches = "\n".join(str(b) for b in branches)
        cset(row, 1, item.get("function", ""))
        cset(row, 2, item.get("decision_point", ""))
        cset(row, 3, item.get("line", ""))
        cset(row, 4, branches)
        row += 1

    # 功能 ↔ 用例（研发核对每个功能的验证用例）
    if cases:
        by_func = {}
        for c in cases:
            by_func.setdefault(c.get("function", "?"), []).append(c.get("case_id", ""))
        if by_func:
            row += 1
            write_section_title(ws, row, "功能 ↔ 用例（每个功能被哪些用例验证）")
            row += 1
            write_table_header(ws, row, ["功能", "用例编号"])
            row += 1
            for fn in sorted(by_func):
                cset(row, 1, fn)
                cset(row, 2, "\n".join(by_func[fn]))
                row += 1

    # ---- 标注区（缺口 / 无法验证 / 规则缺口 合并） ----
    row += 1
    write_section_title(ws, row, "标注区（缺口 / 无法静态验证 / 规则缺口）")
    row += 1
    items = []
    for u in (coverage or {}).get("uncovered") or []:
        items.append(f"【缺口】{u}")
    for n in (coverage or {}).get("notes") or []:
        items.append(f"【无法验证】{n}")
    for g in (coverage or {}).get("verification_gaps") or []:
        items.append(f"【验证缺口】{g}")
    for g in (coverage or {}).get("rule_gaps") or []:
        if isinstance(g, dict):
            g = f"{g.get('category', '')}: {g.get('gap', '')}（{g.get('function', '')}，建议确认：{g.get('confirm_with', '')}）"
        items.append(f"【规则缺口】{g}")
    if not items:
        cset(row, 1, "无（代码可表达范围内已全覆盖，无额外标注）")
        row += 1
    else:
        for it in items:
            cset(row, 1, it)
            row += 1

    # 列宽 + 行高自适应（保证换行内容完全显示）
    ws.column_dimensions["A"].width = COVERAGE_WIDTH_MAP[1]
    ws.column_dimensions["B"].width = COVERAGE_WIDTH_MAP[2]
    ws.column_dimensions["C"].width = COVERAGE_WIDTH_MAP[3]
    ws.column_dimensions["D"].width = COVERAGE_WIDTH_MAP[4]
    for r in range(1, ws.max_row + 1):
        fit_row_height(ws, r, COVERAGE_WIDTH_MAP)
    return ws


def to_cell_value(value):
    if isinstance(value, (list, tuple)):
        return "\n".join(str(v) for v in value)
    if value is None:
        return ""
    return value


def build_workbook(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "研发自测用例"

    # 表头
    for col, (_, title) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = Font(name=FONT_NAME, size=HEADER_FONT_SIZE, bold=True, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # 数据行（斑马纹 + 场景类型整行着色 + 优先级着色 + 行高自适应）
    for row, case in enumerate(cases, start=2):
        font_color = SCENARIO_TYPE_COLORS.get(case.get("scenario_type"))
        band_fill = BAND_FILL_EVEN if row % 2 == 0 else BAND_FILL_ODD
        priority = str(case.get("priority") or "").strip().upper()
        pri_color = PRIORITY_COLORS.get(priority)
        for col, (key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=col, value=to_cell_value(case.get(key)))
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            cell.fill = band_fill
            if key == "priority" and pri_color:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=pri_color)
            elif font_color:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE, color=font_color)
            else:
                cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
        fit_row_height(ws, row, MAIN_WIDTH_MAP)

    # 列宽 + 冻结首行 + 筛选
    for col, (key, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS.get(key, 20)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 测试结果列：PASS/FAIL 下拉（数据验证），允许留空，预留给后续新增行
    test_col = COLUMNS.index(("test_result", "测试结果")) + 1
    test_letter = get_column_letter(test_col)
    dv = DataValidation(
        type="list",
        formula1='"PASS,FAIL"',
        allow_blank=True,
        showDropDown=False,  # False = 显示下拉箭头（Excel 内部 XML 属性是反的）
    )
    dv.error = "只能选择 PASS 或 FAIL"
    dv.errorTitle = "无效输入"
    ws.add_data_validation(dv)
    dv.add(f"{test_letter}2:{test_letter}{max(ws.max_row, 200)}")

    # 场景类型列：正常流程/异常流程 下拉（数据验证）
    st_col = COLUMNS.index(("scenario_type", "场景类型")) + 1
    st_letter = get_column_letter(st_col)
    st_dv = DataValidation(
        type="list",
        formula1='"正常流程,异常流程"',
        allow_blank=True,
        showDropDown=False,
    )
    st_dv.error = "只能选择 正常流程 或 异常流程"
    st_dv.errorTitle = "无效输入"
    ws.add_data_validation(st_dv)
    st_dv.add(f"{st_letter}2:{st_letter}{max(ws.max_row, 200)}")

    return wb


def safe_name(name):
    """清理文件/目录名中的非法字符（Windows 不允许 \\/:*?\"<>|）。"""
    return re.sub(r'[\\/:*?"<>|]', "_", str(name)).strip() or "testcase"


def main():
    parser = argparse.ArgumentParser(description="Generate test-case Excel from JSON")
    parser.add_argument("input", help="Path to JSON file (array of cases or {\"cases\":[...]})")
    parser.add_argument("output", nargs="?", help="Output .xlsx path（指定后覆盖自动命名）")
    parser.add_argument("--project-name", default=None, help="项目名称，用于输出目录与文件名")
    parser.add_argument("--version", default="v1.0", help="版本号，默认 v1.0")
    parser.add_argument("--output-dir", default="testcase", help="输出根目录（默认 testcase），其下按项目名建子目录")
    args = parser.parse_args()

    data = load_data(args.input)
    if isinstance(data, dict) and "cases" in data:
        cases, coverage = data["cases"], data.get("coverage")
    else:
        cases, coverage = data, None

    if args.output:
        out = args.output
    else:
        project = safe_name(args.project_name or "研发自测用例")
        version = str(args.version)
        if not version.startswith("v"):
            version = f"v{version}"
        fname = f"{project}_{date.today():%Y%m%d}_{version}.xlsx"
        out_dir = Path(args.output_dir) / project
        out_dir.mkdir(parents=True, exist_ok=True)
        out = str(out_dir / fname)

    wb = build_workbook(cases)
    if coverage:
        build_coverage_sheet(wb, coverage, cases)
    wb.save(out)
    extra = "（含「覆盖说明」sheet）" if coverage else ""
    print(f"已生成 {len(cases)} 条用例{extra} -> {out}")


if __name__ == "__main__":
    main()
